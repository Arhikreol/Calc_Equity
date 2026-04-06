from __future__ import annotations

import argparse
import sys
from datetime import datetime
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Callable

if "--self-test" in sys.argv:
    import matplotlib

    matplotlib.use("Agg")

from matplotlib import pyplot as plt
from matplotlib.backend_bases import MouseEvent
from matplotlib.widgets import Button, Slider, TextBox
from openpyxl import Workbook, load_workbook

from analyze_group_sequence_money_management import (
    DEFAULT_BASE_STAKE,
    DEFAULT_LOSS_MULT,
    DEFAULT_PAYOUT_PCT,
    DEFAULT_RESET_AFTER_LOSSES,
    DEFAULT_SEQUENCE_PATH,
    SimulationConfig,
    SimulationMetrics,
    SequenceSummary,
    load_sequence,
    run_self_tests as run_analyzer_self_tests,
    simulate_runtime_group_trace,
    summarize_runtime_group_trace,
    summarize_sequence,
)

import tkinter as tk
from tkinter import filedialog

RESET_AFTER_LOSSES_MIN = 2
RESET_AFTER_LOSSES_MAX = 10
LOSS_MULT_MIN = 1.0
LOSS_MULT_MAX = 10.0
LOSS_MULT_STEP = 0.1


@dataclass(frozen=True)
class PlotPayload:
    tradeIndices: list[int]
    equityValues: list[float]
    metrics: SimulationMetrics
    sequenceSummary: SequenceSummary


@dataclass(frozen=True)
class ViewerHandles:
    figure: Any
    plotAxes: Any
    equityLine: Any
    statsText: Any
    autoStatusText: Any
    sourceText: Any
    resetSlider: Slider
    multSlider: Slider
    maxDrawdownBox: TextBox
    loadXlsxButton: Button
    autoBestButton: Button
    applyAutoBest: Callable[[], SimulationConfig | None]
    applyLoadedHistory: Callable[[Path], list[int]]


_ACTIVE_VIEWERS: list[ViewerHandles] = []


@dataclass(frozen=True)
class SequenceSource:
    path: Path
    values: list[int]
    summary: SequenceSummary
    label: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Plot a martingale equity curve for otch.txt with interactive sliders."
    )
    parser.add_argument("--sequence", type=Path, default=DEFAULT_SEQUENCE_PATH)
    parser.add_argument("--reset-after-losses", type=int, default=DEFAULT_RESET_AFTER_LOSSES)
    parser.add_argument("--loss-mult", type=float, default=DEFAULT_LOSS_MULT)
    parser.add_argument("--self-test", action="store_true", help="Run internal checks and exit.")
    return parser.parse_args()


def normalize_reset_after_losses(value: float) -> int:
    return max(RESET_AFTER_LOSSES_MIN, min(RESET_AFTER_LOSSES_MAX, int(round(float(value)))))


def normalize_loss_mult(value: float) -> float:
    return max(LOSS_MULT_MIN, min(LOSS_MULT_MAX, round(float(value) + 1e-12, 1)))


def build_loss_mult_candidates() -> list[float]:
    steps = int(round((LOSS_MULT_MAX - LOSS_MULT_MIN) / LOSS_MULT_STEP))
    return [round(LOSS_MULT_MIN + LOSS_MULT_STEP * index, 1) for index in range(steps + 1)]


def normalize_datetime_sort_key(value: object) -> tuple[int, datetime | str]:
    if isinstance(value, datetime):
        return (0, value)
    if value is None:
        return (2, "")
    text = str(value).strip()
    if not text:
        return (2, "")
    try:
        return (0, datetime.fromisoformat(text))
    except ValueError:
        return (1, text)


def parse_history_row_result(stake_value: object, profit_value: object) -> int | None:
    if profit_value is None:
        return None
    profit = float(profit_value)
    stake = None if stake_value is None else float(stake_value)
    if profit > 0 and stake is not None and profit != stake:
        return 1
    if profit < 0:
        return -1
    return None


def build_sequence_from_history_records(records: list[tuple[object, object, object]]) -> list[int]:
    sorted_records = sorted(records, key=lambda record: normalize_datetime_sort_key(record[0]))
    values: list[int] = []
    for _open_time, stake_value, profit_value in sorted_records:
        result = parse_history_row_result(stake_value, profit_value)
        if result is not None:
            values.append(result)
    if not values:
        raise ValueError("No qualifying trades found in history file")
    return values


def load_history_sequence_from_xlsx(path: Path) -> SequenceSource:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        records: list[tuple[object, object, object]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= 9:
                continue
            records.append((row[4], row[8], row[9]))
    finally:
        workbook.close()

    values = build_sequence_from_history_records(records)
    return SequenceSource(
        path=path,
        values=values,
        summary=summarize_sequence(values),
        label=path.name,
    )


def load_sequence_source(path: Path) -> SequenceSource:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_history_sequence_from_xlsx(path)
    values = load_sequence(path)
    return SequenceSource(
        path=path,
        values=values,
        summary=summarize_sequence(values),
        label=path.name,
    )


def parse_max_drawdown_limit(text: str) -> float | None:
    cleaned = text.strip()
    if not cleaned:
        return None
    value = float(cleaned.replace(",", "."))
    if value < 0:
        raise ValueError("Max DD cap must be >= 0")
    return round(value + 1e-12, 2)


def build_config(reset_after_losses: int, loss_mult: float) -> SimulationConfig:
    return SimulationConfig(
        baseStake=DEFAULT_BASE_STAKE,
        lossMult=normalize_loss_mult(loss_mult),
        resetAfterLosses=normalize_reset_after_losses(reset_after_losses),
        payoutPct=DEFAULT_PAYOUT_PCT,
    )


def build_plot_payload(
    values: list[int],
    config: SimulationConfig,
    sequence_summary: SequenceSummary,
) -> PlotPayload:
    trace = simulate_runtime_group_trace(values, config)
    metrics = summarize_runtime_group_trace(trace)
    trade_indices = [0]
    equity_values = [0.0]
    trade_indices.extend(point.index for point in trace)
    equity_values.extend(point.equity for point in trace)
    return PlotPayload(
        tradeIndices=trade_indices,
        equityValues=equity_values,
        metrics=metrics,
        sequenceSummary=sequence_summary,
    )


def compute_profit_drawdown_ratio(payload: PlotPayload) -> float:
    max_drawdown = payload.metrics.maxDrawdownNormalized
    final_profit = payload.metrics.finalPnlNormalized
    if max_drawdown <= 0:
        if final_profit > 0:
            return float("inf")
        if final_profit < 0:
            return float("-inf")
        return 0.0
    return final_profit / max_drawdown


def find_best_profit_drawdown_config(
    values: list[int],
    sequence_summary: SequenceSummary,
    reset_candidates: list[int] | None = None,
    mult_candidates: list[float] | None = None,
    max_drawdown_limit: float | None = None,
) -> tuple[SimulationConfig, PlotPayload]:
    reset_values = reset_candidates or list(range(RESET_AFTER_LOSSES_MIN, RESET_AFTER_LOSSES_MAX + 1))
    mult_values = mult_candidates or build_loss_mult_candidates()

    best_config: SimulationConfig | None = None
    best_payload: PlotPayload | None = None
    best_rank: tuple[float, float, float, float, float, float] | None = None

    for reset_after_losses in reset_values:
        for loss_mult in mult_values:
            config = build_config(reset_after_losses, loss_mult)
            payload = build_plot_payload(values, config, sequence_summary)
            if (
                max_drawdown_limit is not None
                and payload.metrics.maxDrawdownNormalized > max_drawdown_limit
            ):
                continue
            rank = (
                compute_profit_drawdown_ratio(payload),
                payload.metrics.finalPnlNormalized,
                -payload.metrics.maxDrawdownNormalized,
                -payload.metrics.peakStakeNormalized,
                -config.lossMult,
                -float(config.resetAfterLosses),
            )
            if best_rank is None or rank > best_rank:
                best_config = config
                best_payload = payload
                best_rank = rank

    if best_config is None or best_payload is None:
        if max_drawdown_limit is None:
            raise ValueError("No martingale candidates were produced for auto-search")
        raise ValueError(f"No config fits Max DD <= {max_drawdown_limit:.2f}")
    return best_config, best_payload


def format_stats_text(config: SimulationConfig, payload: PlotPayload) -> str:
    summary = payload.sequenceSummary
    win_rate_pct = 0.0 if summary.tradeCount <= 0 else summary.winCount * 100.0 / summary.tradeCount
    return "\n".join(
        [
            f"reset: {config.resetAfterLosses}",
            f"mult:  {config.lossMult:.1f}x",
            f"ratio: {payload.metrics.recoveryFactor:.4f}",
            f"final: {payload.metrics.finalPnlNormalized:.2f} USD",
            f"max dd:{payload.metrics.maxDrawdownNormalized:.2f} USD",
            f"peak:  {payload.metrics.peakStakeNormalized:.2f} USD",
            f"trades:{summary.tradeCount}",
            f"wins:  {summary.winCount}",
            f"losses:{summary.lossCount}",
            f"wr:    {win_rate_pct:.2f}%",
            f"max ls:{summary.maxLossStreak}",
        ]
    )


def register_active_viewer(handles: ViewerHandles) -> ViewerHandles:
    # Keep widget objects strongly referenced so Button/Slider callbacks survive while the window is open.
    _ACTIVE_VIEWERS[:] = [handles]
    return handles


def trigger_button_click(handles: ViewerHandles) -> None:
    center_x, center_y = handles.autoBestButton.ax.transAxes.transform((0.5, 0.5))
    press_event = MouseEvent("button_press_event", handles.figure.canvas, center_x, center_y, button=1)
    release_event = MouseEvent("button_release_event", handles.figure.canvas, center_x, center_y, button=1)
    handles.figure.canvas.callbacks.process("button_press_event", press_event)
    handles.figure.canvas.callbacks.process("button_release_event", release_event)


def resolve_dialog_parent_window(figure: Any) -> Any | None:
    manager = getattr(getattr(figure, "canvas", None), "manager", None)
    return getattr(manager, "window", None)


def choose_history_xlsx_file(
    figure: Any,
    initial_dir: Path | None = None,
    dialog_fn: Callable[..., str] | None = None,
) -> Path | None:
    dialog = dialog_fn or filedialog.askopenfilename
    parent_window = resolve_dialog_parent_window(figure)
    transient_root: tk.Tk | None = None
    dialog_kwargs: dict[str, object] = {
        "title": "Choose Pocket Option history XLSX",
        "initialdir": str(initial_dir) if initial_dir else None,
        "filetypes": [("Excel files", "*.xlsx"), ("All files", "*.*")],
    }
    if parent_window is not None:
        dialog_kwargs["parent"] = parent_window
    else:
        transient_root = tk.Tk()
        transient_root.withdraw()
        transient_root.attributes("-topmost", True)
        dialog_kwargs["parent"] = transient_root
    try:
        selected = dialog(**dialog_kwargs)
    finally:
        if transient_root is not None:
            transient_root.destroy()
    if not selected:
        return None
    return Path(selected)


def create_viewer(sequence_path: Path, values: list[int], initial_config: SimulationConfig):
    source = SequenceSource(
        path=sequence_path,
        values=list(values),
        summary=summarize_sequence(values),
        label=sequence_path.name,
    )
    payload = build_plot_payload(source.values, initial_config, source.summary)

    figure = plt.figure(figsize=(13, 7))
    manager = getattr(figure.canvas, "manager", None)
    if manager is not None and hasattr(manager, "set_window_title"):
        manager.set_window_title("OTCH Equity Viewer")

    plot_ax = figure.add_axes([0.08, 0.30, 0.65, 0.62])
    stats_ax = figure.add_axes([0.76, 0.44, 0.20, 0.48])
    source_ax = figure.add_axes([0.76, 0.38, 0.20, 0.05])
    load_ax = figure.add_axes([0.76, 0.32, 0.20, 0.05])
    max_dd_ax = figure.add_axes([0.76, 0.28, 0.10, 0.05])
    auto_ax = figure.add_axes([0.87, 0.27, 0.09, 0.06])
    auto_status_ax = figure.add_axes([0.76, 0.22, 0.20, 0.04])
    reset_ax = figure.add_axes([0.08, 0.16, 0.88, 0.05])
    mult_ax = figure.add_axes([0.08, 0.08, 0.88, 0.05])

    (equity_line,) = plot_ax.plot(
        payload.tradeIndices,
        payload.equityValues,
        color="#1f77b4",
        linewidth=2.0,
    )
    plot_ax.axhline(0.0, color="#808080", linewidth=1.0, linestyle="--")
    plot_ax.set_title(f"Martingale equity: {sequence_path.name}")
    plot_ax.set_xlabel("Trade index")
    plot_ax.set_ylabel("Equity, USD")
    plot_ax.grid(True, alpha=0.3)
    plot_ax.set_xlim(0, payload.tradeIndices[-1])

    stats_ax.axis("off")
    source_ax.axis("off")
    stats_text = stats_ax.text(
        0.0,
        1.0,
        format_stats_text(initial_config, payload),
        va="top",
        ha="left",
        family="monospace",
        fontsize=11,
    )
    source_text = source_ax.text(
        0.0,
        0.8,
        f"Source: {source.label}",
        va="top",
        ha="left",
        fontsize=9,
    )

    reset_slider = Slider(
        ax=reset_ax,
        label="Reset after losses",
        valmin=2.0,
        valmax=10.0,
        valinit=float(initial_config.resetAfterLosses),
        valstep=1.0,
    )
    mult_slider = Slider(
        ax=mult_ax,
        label="Martingale multiplier",
        valmin=LOSS_MULT_MIN,
        valmax=LOSS_MULT_MAX,
        valinit=initial_config.lossMult,
        valstep=LOSS_MULT_STEP,
    )
    load_xlsx_button = Button(load_ax, "Load XLSX")
    max_drawdown_box = TextBox(max_dd_ax, "Max DD", initial="")
    auto_best_button = Button(auto_ax, "Auto best P/DD")
    auto_status_ax.axis("off")
    auto_status_text = auto_status_ax.text(
        0.0,
        0.75,
        "Auto: no DD cap",
        va="top",
        ha="left",
        fontsize=9,
    )

    is_syncing = False

    def render_payload(config: SimulationConfig, updated_payload: PlotPayload, source_label: str) -> None:
        equity_line.set_data(updated_payload.tradeIndices, updated_payload.equityValues)
        plot_ax.relim()
        plot_ax.autoscale_view(scalex=False, scaley=True)
        plot_ax.set_xlim(0, updated_payload.tradeIndices[-1])
        plot_ax.set_title(f"Martingale equity: {source_label}")
        stats_text.set_text(format_stats_text(config, updated_payload))
        source_text.set_text(f"Source: {source_label}")
        figure.canvas.draw_idle()

    def sync_controls(config: SimulationConfig) -> None:
        nonlocal is_syncing
        is_syncing = True
        reset_slider.set_val(float(config.resetAfterLosses))
        mult_slider.set_val(config.lossMult)
        is_syncing = False

    def set_auto_status(message: str, *, color: str = "#333333") -> None:
        auto_status_text.set_text(message)
        auto_status_text.set_color(color)

    def refresh_plot(_value: float) -> None:
        nonlocal source
        if is_syncing:
            return

        config = build_config(reset_slider.val, mult_slider.val)
        if abs(reset_slider.val - config.resetAfterLosses) > 1e-9 or abs(mult_slider.val - config.lossMult) > 1e-9:
            sync_controls(config)
        render_payload(config, build_plot_payload(source.values, config, source.summary), source.label)

    def apply_auto_best() -> SimulationConfig | None:
        try:
            max_drawdown_limit = parse_max_drawdown_limit(max_drawdown_box.text)
            config, best_payload = find_best_profit_drawdown_config(
                source.values,
                source.summary,
                max_drawdown_limit=max_drawdown_limit,
            )
        except ValueError as exc:
            set_auto_status(str(exc), color="#b22222")
            figure.canvas.draw_idle()
            return None
        sync_controls(config)
        render_payload(config, best_payload, source.label)
        if max_drawdown_limit is None:
            set_auto_status("Auto: no DD cap")
        else:
            set_auto_status(f"Auto: DD <= {max_drawdown_limit:.2f}")
        figure.canvas.draw_idle()
        return config

    def apply_loaded_history(path: Path) -> list[int]:
        nonlocal source
        source = load_sequence_source(path)
        config = build_config(reset_slider.val, mult_slider.val)
        render_payload(config, build_plot_payload(source.values, config, source.summary), source.label)
        set_auto_status("Loaded history file")
        return list(source.values)

    def on_auto_best_click(_event: object) -> None:
        apply_auto_best()

    def on_load_xlsx_click(_event: object) -> None:
        selected = choose_history_xlsx_file(figure, source.path.parent)
        if selected is None:
            set_auto_status("Load cancelled")
            figure.canvas.draw_idle()
            return
        try:
            apply_loaded_history(selected)
        except ValueError as exc:
            set_auto_status(str(exc), color="#b22222")
            figure.canvas.draw_idle()

    reset_slider.on_changed(refresh_plot)
    mult_slider.on_changed(refresh_plot)
    load_xlsx_button.on_clicked(on_load_xlsx_click)
    auto_best_button.on_clicked(on_auto_best_click)
    return ViewerHandles(
        figure=figure,
        plotAxes=plot_ax,
        equityLine=equity_line,
        statsText=stats_text,
        autoStatusText=auto_status_text,
        sourceText=source_text,
        resetSlider=reset_slider,
        multSlider=mult_slider,
        maxDrawdownBox=max_drawdown_box,
        loadXlsxButton=load_xlsx_button,
        autoBestButton=auto_best_button,
        applyAutoBest=apply_auto_best,
        applyLoadedHistory=apply_loaded_history,
    )


def run_self_tests() -> None:
    run_analyzer_self_tests()

    config = build_config(3, 3.0)
    four_loss_summary = summarize_sequence([-1, -1, -1, -1])
    payload = build_plot_payload([-1, -1, -1, -1], config, four_loss_summary)
    assert payload.tradeIndices == [0, 1, 2, 3, 4]
    assert payload.equityValues == [0.0, -1.0, -4.0, -13.0, -14.0]
    assert payload.metrics.peakStakeNormalized == 9.0
    assert normalize_reset_after_losses(4.0) == 4
    assert normalize_loss_mult(3.06) == 3.1
    assert build_loss_mult_candidates()[:3] == [1.0, 1.1, 1.2]
    assert parse_max_drawdown_limit("") is None
    assert parse_max_drawdown_limit("96,5") == 96.5
    try:
        parse_max_drawdown_limit("-1")
    except ValueError as exc:
        assert "Max DD cap" in str(exc)
    else:
        raise AssertionError("Negative Max DD cap must fail")

    sorted_values = build_sequence_from_history_records(
        [
            ("2026-04-05 03:30:00", 2.5, -2.5),
            ("2026-04-05 02:30:00", 2.5, 2.3),
            ("2026-04-05 02:56:00", 1.0, -1.0),
            ("2026-04-05 02:40:00", 3.0, 3.0),
            ("2026-04-05 02:20:00", 6.25, 5.75),
        ]
    )
    assert sorted_values == [1, 1, -1, -1]

    with TemporaryDirectory() as temp_dir:
        workbook_path = Path(temp_dir) / "history.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(
            [
                "Direction",
                "Deal",
                "Expiration",
                "Asset",
                "Open time",
                "Close time",
                "Open price",
                "Close price",
                "Stake",
                "Profit",
            ]
        )
        worksheet.append(["put", "a", "S60", "X", "2026-04-05 03:30:00", "", 0, 0, 2.5, -2.5])
        worksheet.append(["put", "b", "S60", "X", "2026-04-05 02:30:00", "", 0, 0, 2.5, 2.3])
        worksheet.append(["put", "c", "S60", "X", "2026-04-05 02:40:00", "", 0, 0, 3.0, 3.0])
        workbook.save(workbook_path)
        workbook.close()

        workbook_source = load_sequence_source(workbook_path)
        assert workbook_source.values == [1, -1]
        assert workbook_source.summary.tradeCount == 2
        assert workbook_source.label == "history.xlsx"

    dummy_parent = object()
    fake_figure = type(
        "FakeFigure",
        (),
        {"canvas": type("FakeCanvas", (), {"manager": type("FakeManager", (), {"window": dummy_parent})()})()},
    )()
    captured_kwargs: dict[str, object] = {}

    def fake_dialog(**kwargs: object) -> str:
        captured_kwargs.update(kwargs)
        return ""

    assert choose_history_xlsx_file(fake_figure, Path("."), dialog_fn=fake_dialog) is None
    assert captured_kwargs["parent"] is dummy_parent

    stats_summary = summarize_sequence([1, -1, 1, -1])
    stats_payload = build_plot_payload([1, -1, 1, -1], config, stats_summary)
    stats_text = format_stats_text(config, stats_payload)
    assert "trades:4" in stats_text
    assert "wins:  2" in stats_text
    assert "losses:2" in stats_text
    assert "wr:    50.00%" in stats_text

    best_config, best_payload = find_best_profit_drawdown_config(
        [-1, -1, 1, 1],
        summarize_sequence([-1, -1, 1, 1]),
        reset_candidates=[2, 3],
        mult_candidates=[1.0, 3.0],
    )
    assert best_config.resetAfterLosses == 3
    assert best_config.lossMult == 3.0
    assert best_payload.metrics.recoveryFactor == 1.3
    capped_config, _capped_payload = find_best_profit_drawdown_config(
        [-1, -1, 1, 1],
        summarize_sequence([-1, -1, 1, 1]),
        reset_candidates=[2, 3],
        mult_candidates=[1.0, 3.0],
        max_drawdown_limit=2.0,
    )
    assert capped_config.resetAfterLosses == 2
    assert capped_config.lossMult == 1.0
    try:
        find_best_profit_drawdown_config(
            [-1, -1, 1, 1],
            summarize_sequence([-1, -1, 1, 1]),
            reset_candidates=[2, 3],
            mult_candidates=[1.0, 3.0],
            max_drawdown_limit=1.0,
        )
    except ValueError as exc:
        assert "No config fits Max DD" in str(exc)
    else:
        raise AssertionError("Auto-search must fail when no config satisfies Max DD limit")

    handles = register_active_viewer(create_viewer(Path("otch.txt"), [-1, -1, -1, 1], config))
    baseline_y = list(handles.equityLine.get_ydata())
    handles.resetSlider.set_val(2.0)
    updated_text = handles.statsText.get_text()
    updated_y = list(handles.equityLine.get_ydata())
    assert "reset: 2" in updated_text
    assert baseline_y != updated_y
    handles.multSlider.set_val(4.0)
    assert "mult:  4.0x" in handles.statsText.get_text()
    best_config = handles.applyAutoBest()
    assert handles.resetSlider.val == float(best_config.resetAfterLosses)
    assert handles.multSlider.val == best_config.lossMult
    assert "ratio:" in handles.statsText.get_text()
    assert "wr:" in handles.statsText.get_text()

    click_handles = register_active_viewer(create_viewer(Path("otch.txt"), [-1, -1, 1, 1], build_config(2, 1.0)))
    click_handles.maxDrawdownBox.set_val("4.0")
    before_click = click_handles.statsText.get_text()
    trigger_button_click(click_handles)
    after_click = click_handles.statsText.get_text()
    assert before_click != after_click
    assert "ratio:" in after_click
    assert click_handles.resetSlider.val == 3.0
    assert click_handles.multSlider.val == 3.0
    assert "DD <= 4.00" in click_handles.autoStatusText.get_text()
    plt.close(click_handles.figure)

    load_handles = register_active_viewer(create_viewer(Path("otch.txt"), [1, -1], build_config(3, 3.0)))
    with TemporaryDirectory() as temp_dir:
        workbook_path = Path(temp_dir) / "history.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"])
        worksheet.append(["", "", "", "", "2026-04-05 03:30:00", "", "", "", 2.5, -2.5])
        worksheet.append(["", "", "", "", "2026-04-05 02:30:00", "", "", "", 2.5, 2.3])
        workbook.save(workbook_path)
        workbook.close()
        loaded_values = load_handles.applyLoadedHistory(workbook_path)
        assert loaded_values == [1, -1]
        assert "history.xlsx" in load_handles.sourceText.get_text()
        assert load_handles.plotAxes.get_title() == "Martingale equity: history.xlsx"
    plt.close(load_handles.figure)
    plt.close(handles.figure)


def main() -> None:
    args = parse_args()
    if args.self_test:
        run_self_tests()
        print("self-test: PASS")
        return

    values = load_sequence(args.sequence)
    config = build_config(args.reset_after_losses, args.loss_mult)
    register_active_viewer(create_viewer(args.sequence, values, config))
    plt.show()


if __name__ == "__main__":
    main()
